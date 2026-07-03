"""
Trainers Excel Import & Export
================================

Companion module to crm/api/trainers.py - adds bulk import/export on top
of the existing CRM Trainer CRUD without touching it.

Flow
----
1. download_template()           -> .xlsx with headers + 2 sample rows
2. parse_and_validate(file_url)  -> reads the uploaded file, validates every
                                     row, returns a preview payload (never
                                     writes to the DB)
3. commit_import(rows, mode)     -> the user has reviewed the preview and
                                     clicks "Import" - this actually writes
                                     CRM Trainer records, in a background
                                     job when the row count is large, and
                                     always finishes by writing a
                                     CRM Trainer Import Log entry
4. export_trainers(...)          -> streams .xlsx/.csv of filtered trainers

Permissions
-----------
Every entry point calls _require_admin() - only System Manager (the
project's existing "admin" role check, same convention as
get_session_role_flags in crm/api/session.py) may import or export.
This is enforced server-side; the frontend also hides the buttons for
non-admins, but that's UX only, not the real gate.

Column mapping (Excel header -> CRM Trainer fieldname)
--------------------------------------------------------
Matches the 13 fields already on CRM Trainer 1:1 - no schema change to
the trainer doctype itself.
"""

import io
import json
import re
from datetime import datetime

import frappe
from frappe import _
from frappe.utils import now_datetime, cint, flt

from crm.api.session import get_session_role_flags


# ─────────────────────────────────────────────────────────────────────────
#  Column definitions - single source of truth for template, parsing,
#  validation, and export so the three never drift out of sync.
# ─────────────────────────────────────────────────────────────────────────

COLUMNS = [
	{"header": "Trainer Name *", "field": "trainer_name", "required": True, "type": "text"},
	{"header": "Phone Number", "field": "phone", "required": False, "type": "phone"},
	{"header": "Alternative Phone Number", "field": "alternate_phone", "required": False, "type": "phone"},
	{"header": "Email", "field": "email", "required": False, "type": "email"},
	{"header": "LinkedIn Profile", "field": "linkedin_profile", "required": False, "type": "text"},
	{"header": "Location", "field": "location", "required": False, "type": "text"},
	{"header": "Technology Expert In", "field": "technology_expert_in", "required": False, "type": "text"},
	{"header": "Skill Level", "field": "skill_level", "required": False, "type": "select",
	 "options": ["Beginner", "Intermediate", "Advanced", "Expert"]},
	{"header": "Experience", "field": "experience", "required": False, "type": "text"},
	{"header": "Availability", "field": "availability", "required": False, "type": "select",
	 "options": ["Available", "Partially Available", "Not Available"]},
	{"header": "Status", "field": "status", "required": False, "type": "select",
	 "options": ["Active", "Inactive", "Blacklisted"]},
	{"header": "Commercial (₹/day)", "field": "commercial", "required": False, "type": "number"},
	{"header": "Commercial Type", "field": "commercial_type", "required": False, "type": "select",
	 "options": ["Per Day", "Per Hour"]},
	{"header": "Company", "field": "company", "required": False, "type": "text"},
	{"header": "Remarks", "field": "remarks", "required": False, "type": "text"},
]

FIELD_BY_HEADER = {c["header"]: c["field"] for c in COLUMNS}
# Tolerant header matching: also accept the bare field-style header without
# the "*" or unit suffix, in case a user edits the template's header row.
LOOSE_HEADER_MAP = {
	"trainer name": "trainer_name",
	"phone number": "phone",
	"alternative phone number": "alternate_phone",
	"alternate phone number": "alternate_phone",
	"email": "email",
	"linkedin profile": "linkedin_profile",
	"location": "location",
	"technology expert in": "technology_expert_in",
	"skill level": "skill_level",
	"experience": "experience",
	"availability": "availability",
	"status": "status",
	"commercial": "commercial",
	"commercial type": "commercial_type",
	"company": "company",
	"remarks": "remarks",
}

SAMPLE_ROWS = [
	{
		"trainer_name": "Ravi Kumar", "phone": "+91 9876543210", "email": "ravi.kumar@example.com",
		"linkedin_profile": "https://linkedin.com/in/ravikumar", "location": "Bangalore",
		"technology_expert_in": "Python, Django", "skill_level": "Expert", "experience": "8 years",
		"availability": "Available", "status": "Active", "commercial": 15000,
		"company": "Freelance", "remarks": "Strong in backend training",
	},
	{
		"trainer_name": "Sneha Reddy", "phone": "9123456780", "email": "sneha.reddy@example.com",
		"linkedin_profile": "https://linkedin.com/in/snehareddy", "location": "Hyderabad",
		"technology_expert_in": "React, JavaScript", "skill_level": "Advanced", "experience": "5 years",
		"availability": "Partially Available", "status": "Active", "commercial": 12000,
		"company": "TechMentors", "remarks": "",
	},
]

MAX_PREVIEW_ROWS = 5000
SYNC_THRESHOLD = 500  # rows; above this, commit_import runs as a background job


# ─────────────────────────────────────────────────────────────────────────
#  Access control
# ─────────────────────────────────────────────────────────────────────────

def _require_admin():
	roles = get_session_role_flags()
	if not roles.get("is_system_manager"):
		frappe.throw(_("Only System Manager can import or export Trainers."), frappe.PermissionError)


# ─────────────────────────────────────────────────────────────────────────
#  1. Download Template
# ─────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def download_template():
	"""Streams an .xlsx with the 13 trainer columns + 2 sample rows."""
	_require_admin()

	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill

	wb = Workbook()
	sheet = wb.active
	sheet.title = "Trainers"

	headers = [c["header"] for c in COLUMNS]
	sheet.append(headers)

	header_fill = PatternFill("solid", start_color="DBEAFE")
	for cell in sheet[1]:
		cell.font = Font(bold=True)
		cell.fill = header_fill

	for sample in SAMPLE_ROWS:
		sheet.append([sample.get(c["field"], "") for c in COLUMNS])

	for col in sheet.columns:
		max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
		sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

	buffer = io.BytesIO()
	wb.save(buffer)

	frappe.response["filename"] = "trainer_import_template.xlsx"
	frappe.response["filecontent"] = buffer.getvalue()
	frappe.response["type"] = "download"


# ─────────────────────────────────────────────────────────────────────────
#  2. Parse + Validate (preview - never writes to DB)
# ─────────────────────────────────────────────────────────────────────────

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
# Accepts +91XXXXXXXXXX, 10-digit Indian mobile, or a generic 7-15 digit
# international number with optional +, spaces, hyphens.
PHONE_RE = re.compile(r"^\+?[\d\s\-]{7,17}$")


@frappe.whitelist()
def parse_and_validate(file_url=None):
	"""
	Reads an already-uploaded file (file_url from Frappe's File doctype,
	populated by the frontend's drag-and-drop uploader) and returns a
	preview payload: every row plus per-row validation errors, and
	duplicate detection against existing CRM Trainer records by email/phone
	AND within the file itself.

	Returns
	-------
	{
	  "total_rows": int,
	  "valid_rows": int,
	  "invalid_rows": int,
	  "duplicate_rows": int,
	  "rows": [
	    {
	      "row_number": 2,                 # 1-indexed, matches Excel row (header = row 1)
	      "data": {...mapped fields...},
	      "errors": ["Email is invalid"],  # empty list = no errors
	      "duplicate_of": "TRAINER-0001" or None,
	      "duplicate_in_file": True/False,
	    },
	    ...
	  ]
	}
	"""
	_require_admin()

	if not file_url:
		frappe.throw(_("file_url is required"))

	raw_rows, parse_error = _read_uploaded_file(file_url)
	if parse_error:
		frappe.throw(parse_error)

	if len(raw_rows) > MAX_PREVIEW_ROWS:
		frappe.throw(
			_("File has {0} rows, which exceeds the {1} row limit per import. Please split the file.").format(
				len(raw_rows), MAX_PREVIEW_ROWS
			)
		)

	# Pre-fetch existing trainers' email/phone for duplicate detection in
	# one query instead of one query per row (avoids N+1).
	existing = frappe.get_all(
		"CRM Trainer",
		fields=["name", "email", "phone"],
		ignore_permissions=True,
	)
	existing_by_email = {e["email"].strip().lower(): e["name"] for e in existing if e.get("email")}
	existing_by_phone = {_normalize_phone(e["phone"]): e["name"] for e in existing if e.get("phone")}

	seen_emails_in_file = {}
	seen_phones_in_file = {}

	result_rows = []
	valid_count = 0
	invalid_count = 0
	duplicate_count = 0

	for idx, raw in enumerate(raw_rows):
		row_number = idx + 2  # +1 for 0-index, +1 because row 1 is the header
		mapped = _map_row(raw)
		errors = _validate_row(mapped, row_number)

		email_key = mapped.get("email", "").strip().lower() if mapped.get("email") else None
		phone_key = _normalize_phone(mapped.get("phone")) if mapped.get("phone") else None

		duplicate_of = None
		duplicate_in_file = False

		if email_key and email_key in existing_by_email:
			duplicate_of = existing_by_email[email_key]
		elif phone_key and phone_key in existing_by_phone:
			duplicate_of = existing_by_phone[phone_key]

		if email_key and email_key in seen_emails_in_file:
			duplicate_in_file = True
		if phone_key and phone_key in seen_phones_in_file:
			duplicate_in_file = True

		if email_key:
			seen_emails_in_file.setdefault(email_key, row_number)
		if phone_key:
			seen_phones_in_file.setdefault(phone_key, row_number)

		if errors:
			invalid_count += 1
		elif duplicate_of or duplicate_in_file:
			duplicate_count += 1
			valid_count += 1  # duplicates are still structurally valid rows
		else:
			valid_count += 1

		result_rows.append({
			"row_number": row_number,
			"data": mapped,
			"errors": errors,
			"duplicate_of": duplicate_of,
			"duplicate_in_file": duplicate_in_file,
		})

	return {
		"total_rows": len(result_rows),
		"valid_rows": valid_count,
		"invalid_rows": invalid_count,
		"duplicate_rows": duplicate_count,
		"rows": result_rows,
	}


def _read_uploaded_file(file_url):
	"""
	Returns (list_of_row_dicts, error_message). Supports .xlsx and .csv.
	Each row dict has RAW header strings as keys (not yet mapped to
	fieldnames) - that happens in _map_row so we can tolerate header drift.
	"""
	file_doc = frappe.get_doc("File", {"file_url": file_url})
	file_path = file_doc.get_full_path()
	filename = (file_doc.file_name or "").lower()

	try:
		if filename.endswith(".csv"):
			return _read_csv(file_path), None
		elif filename.endswith(".xlsx") or filename.endswith(".xls"):
			return _read_xlsx(file_path), None
		else:
			return [], _("Unsupported file type. Please upload a .xlsx or .csv file.")
	except Exception:
		frappe.log_error(title="Trainer Import - File Parse Error", message=frappe.get_traceback())
		return [], _("Could not read the uploaded file. Please check it isn't corrupted and try again.")


def _read_xlsx(file_path):
	from openpyxl import load_workbook

	wb = load_workbook(file_path, read_only=True, data_only=True)
	sheet = wb.active

	rows_iter = sheet.iter_rows(values_only=True)
	header_row = next(rows_iter, None)
	if not header_row:
		return []

	headers = [str(h).strip() if h is not None else "" for h in header_row]

	rows = []
	for raw in rows_iter:
		if raw is None or all(v is None for v in raw):
			continue  # skip fully blank rows
		row_dict = {}
		for i, header in enumerate(headers):
			if not header:
				continue
			value = raw[i] if i < len(raw) else None
			row_dict[header] = "" if value is None else str(value).strip()
		rows.append(row_dict)

	return rows


def _read_csv(file_path):
	import csv

	rows = []
	with open(file_path, newline="", encoding="utf-8-sig") as f:
		reader = csv.DictReader(f)
		for raw in reader:
			if not any((v or "").strip() for v in raw.values()):
				continue
			rows.append({k.strip(): (v or "").strip() for k, v in raw.items()})
	return rows


def _map_row(raw_row):
	"""Map a raw {header: value} dict to {fieldname: value} using both the
	exact template headers and a tolerant lowercase fallback."""
	mapped = {}
	for header, value in raw_row.items():
		field = FIELD_BY_HEADER.get(header)
		if not field:
			field = LOOSE_HEADER_MAP.get(header.strip().lower().rstrip("*").strip())
		if field:
			mapped[field] = value
	return mapped


def _normalize_phone(phone):
	if not phone:
		return None
	digits = re.sub(r"\D", "", str(phone))
	# Normalize Indian numbers with country code to bare 10-digit form so
	# "+91 98765 43210" and "9876543210" are recognised as the same number.
	if len(digits) == 12 and digits.startswith("91"):
		digits = digits[2:]
	return digits or None


def _validate_row(mapped, row_number):
	"""
	Trainer Name is the ONLY mandatory field. Every other field is
	validated only when a value was actually provided - an empty/blank
	cell for any optional field never produces an error, regardless of
	what that field's format rules would otherwise require.
	"""
	errors = []

	trainer_name = (mapped.get("trainer_name") or "").strip()
	if not trainer_name:
		errors.append(_("Row {0}: Trainer Name is required").format(row_number))

	email = (mapped.get("email") or "").strip()
	if email and not EMAIL_RE.match(email):
		errors.append(_("Row {0}: Email '{1}' is not valid").format(row_number, email))

	phone = (mapped.get("phone") or "").strip()
	if phone and not PHONE_RE.match(phone):
		errors.append(_("Row {0}: Phone number '{1}' is not valid").format(row_number, phone))

	commercial = (mapped.get("commercial") or "")
	commercial = commercial.strip() if isinstance(commercial, str) else commercial
	if commercial not in (None, ""):
		try:
			flt(commercial)
		except (TypeError, ValueError):
			errors.append(_("Row {0}: Commercial value '{1}' must be numeric").format(row_number, commercial))

	skill_level = (mapped.get("skill_level") or "").strip()
	if skill_level and skill_level not in ["Beginner", "Intermediate", "Advanced", "Expert"]:
		errors.append(_("Row {0}: Skill Level '{1}' is not a valid option").format(row_number, skill_level))

	availability = (mapped.get("availability") or "").strip()
	if availability and availability not in ["Available", "Partially Available", "Not Available"]:
		errors.append(_("Row {0}: Availability '{1}' is not a valid option").format(row_number, availability))

	status = (mapped.get("status") or "").strip()
	if status and status not in ["Active", "Inactive", "Blacklisted"]:
		errors.append(_("Row {0}: Status '{1}' is not a valid option").format(row_number, status))

	return errors


# ─────────────────────────────────────────────────────────────────────────
#  3. Commit Import
# ─────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def commit_import(rows, mode="skip", file_name="import.xlsx"):
	"""
	Actually writes CRM Trainer records from previously-validated rows.

	Parameters
	----------
	rows : JSON string or list of the SAME row dicts returned by
	       parse_and_validate (each must have "data", "errors",
	       "duplicate_of"). Rows with non-empty "errors" are always
	       skipped regardless of mode - the frontend should already
	       exclude them, but this is enforced again server-side.
	mode : "skip" (Skip Existing Records) | "update" (Update Existing
	       Records) | "create" (Create New Only - duplicates are
	       skipped, never updated)
	file_name : original uploaded filename, for the import log

	For row counts above SYNC_THRESHOLD this enqueues a background job
	and returns immediately with a job reference; the frontend should
	poll get_import_log(name) or just refresh Import History after a
	short delay. For smaller imports it runs synchronously and returns
	the full summary right away.
	"""
	_require_admin()

	if isinstance(rows, str):
		rows = json.loads(rows)

	if mode not in ("skip", "update", "create"):
		frappe.throw(_("Invalid import mode"))

	# Always drop rows with validation errors - belt and suspenders even
	# though the frontend already filters these out before calling commit.
	clean_rows = [r for r in rows if not r.get("errors")]

	if len(clean_rows) > SYNC_THRESHOLD:
		frappe.enqueue(
			"crm.api.trainers_import._run_import",
			queue="long",
			job_name=f"trainer_import_{frappe.session.user}_{now_datetime()}",
			rows=clean_rows,
			mode=mode,
			file_name=file_name,
			user=frappe.session.user,
		)
		return {
			"queued": True,
			"message": _("Import of {0} records has been queued and is running in the background. "
						  "Check Import History shortly for results.").format(len(clean_rows)),
		}

	summary = _run_import(clean_rows, mode, file_name, frappe.session.user)
	summary["queued"] = False
	return summary


def _run_import(rows, mode, file_name, user):
	"""
	The actual write loop. Safe to call both synchronously and from
	frappe.enqueue (hence the explicit `user` param instead of relying on
	frappe.session.user, since background jobs run in their own context).
	"""
	success_count = 0
	updated_count = 0
	skipped_count = 0
	failed_count = 0
	error_details = []

	for row in rows:
		data = row.get("data", {})
		duplicate_of = row.get("duplicate_of")
		row_number = row.get("row_number")

		try:
			if duplicate_of:
				if mode == "skip":
					skipped_count += 1
					continue
				elif mode == "create":
					skipped_count += 1
					continue
				elif mode == "update":
					doc = frappe.get_doc("CRM Trainer", duplicate_of)
					doc.update(_clean_for_doc(data, is_update=True))
					doc.save(ignore_permissions=True)
					updated_count += 1
			else:
				doc = frappe.new_doc("CRM Trainer")
				doc.update(_clean_for_doc(data, is_update=False))
				doc.insert(ignore_permissions=True)
				success_count += 1

		except Exception as e:
			failed_count += 1
			error_details.append({
				"row": row_number,
				"trainer_name": data.get("trainer_name"),
				"message": str(e),
			})
			frappe.log_error(
				title="Trainer Import Row Failed",
				message=f"Row {row_number}: {frappe.get_traceback()}",
			)

	frappe.db.commit()

	log = frappe.new_doc("CRM Trainer Import Log")
	log.file_name = file_name
	log.imported_by = user
	log.import_datetime = now_datetime()
	log.total_records = len(rows)
	log.success_count = success_count
	log.updated_count = updated_count
	log.skipped_count = skipped_count
	log.failed_count = failed_count
	log.import_mode = {"skip": "Skip Existing", "update": "Update Existing", "create": "Create New Only"}[mode]
	log.error_details = json.dumps(error_details, indent=2) if error_details else ""
	log.insert(ignore_permissions=True)
	frappe.db.commit()

	return {
		"total_rows": len(rows),
		"success_count": success_count,
		"updated_count": updated_count,
		"skipped_count": skipped_count,
		"failed_count": failed_count,
		"log_name": log.name,
	}


def _clean_for_doc(data, is_update=False):
	"""
	Build the dict passed to doc.update() before insert/save.

	Trainer Name is the only field that must have a real value (already
	enforced in _validate_row - this function assumes the row already
	passed validation).

	Two different behaviors depending on is_update:

	- is_update=False (new trainer, doc.insert()):
	  every optional field is included explicitly, blank cells become
	  None. Frappe stores None as a proper NULL/empty value - this is
	  what makes "empty cells are stored as null/empty in the database"
	  true for newly imported trainers, rather than silently falling
	  back to a field's schema default (e.g. Status defaulting to
	  "Active" when the cell was actually left blank on purpose).

	- is_update=True (existing trainer, doc.save() after re-matching by
	  email/phone): blank cells are OMITTED instead of sent as None, so
	  re-importing a file that only has a few columns filled in for an
	  existing trainer does not wipe out fields that trainer already
	  had data in. Only cells that actually have a value in the Excel
	  row will overwrite the existing record.

	commercial (Currency) is the one numeric field: blank becomes None
	(insert) or is skipped (update); a value that still can't be parsed
	at this point (should never happen, since _validate_row already
	rejected it) is also treated as blank rather than crashing the insert.
	"""
	cleaned = {}
	for field, value in data.items():
		if field == "trainer_name":
			cleaned[field] = (value or "").strip()
			continue

		is_blank = value is None or (isinstance(value, str) and value.strip() == "")

		if is_blank:
			if is_update:
				continue  # leave the existing value untouched
			cleaned[field] = None
			continue

		if field == "commercial":
			try:
				cleaned[field] = flt(value)
			except (TypeError, ValueError):
				if not is_update:
					cleaned[field] = None
				continue
		else:
			cleaned[field] = value.strip() if isinstance(value, str) else value

	return cleaned


@frappe.whitelist()
def get_import_logs(limit=20):
	"""Import History list - file name, imported by, timestamp, counts."""
	_require_admin()
	return frappe.get_all(
		"CRM Trainer Import Log",
		fields=[
			"name", "file_name", "imported_by", "import_datetime",
			"total_records", "success_count", "updated_count",
			"skipped_count", "failed_count", "import_mode",
		],
		order_by="import_datetime desc",
		limit=cint(limit) or 20,
		ignore_permissions=True,
	)


@frappe.whitelist()
def get_import_log_detail(name):
	"""Full detail of one import log, including parsed error_details."""
	_require_admin()
	doc = frappe.get_doc("CRM Trainer Import Log", name)
	result = doc.as_dict()
	try:
		result["error_details"] = json.loads(doc.error_details) if doc.error_details else []
	except (TypeError, ValueError):
		result["error_details"] = []
	return result


# ─────────────────────────────────────────────────────────────────────────
#  4. Export Trainers
# ─────────────────────────────────────────────────────────────────────────

EXPORT_PRESETS = {
	"all": {},
	"active": {"status": "Active"},
	"available": {"availability": "Available"},
}


@frappe.whitelist()
def export_trainers(preset="all", filters=None, format="xlsx"):
	"""
	Streams a file download of trainers matching either a named preset
	(all / active / available) or an explicit filters dict (for "Filtered
	Trainers" - whatever filter state the Trainers list view currently has).
	"""
	_require_admin()

	if filters:
		if isinstance(filters, str):
			filters = json.loads(filters)
		_filters = filters
	else:
		_filters = EXPORT_PRESETS.get(preset, {})

	trainers = frappe.get_all(
		"CRM Trainer",
		filters=_filters,
		fields=["name"] + [c["field"] for c in COLUMNS],
		order_by="trainer_name asc",
		ignore_permissions=True,
	)

	headers = [c["header"] for c in COLUMNS]
	rows = [[t.get(c["field"], "") for c in COLUMNS] for t in trainers]

	preset_label = preset if not filters else "filtered"
	filename_base = f"trainers-{preset_label}-{now_datetime().strftime('%Y%m%d-%H%M%S')}"

	if format == "csv":
		_export_csv(headers, rows, filename_base)
	else:
		_export_xlsx(headers, rows, filename_base)


def _export_csv(headers, rows, filename_base):
	import csv

	buffer = io.StringIO()
	writer = csv.writer(buffer)
	writer.writerow(headers)
	writer.writerows(rows)

	frappe.response["filename"] = f"{filename_base}.csv"
	frappe.response["filecontent"] = buffer.getvalue()
	frappe.response["type"] = "download"


def _export_xlsx(headers, rows, filename_base):
	from openpyxl import Workbook
	from openpyxl.styles import Font, PatternFill

	wb = Workbook()
	sheet = wb.active
	sheet.title = "Trainers"

	sheet.append(headers)
	header_fill = PatternFill("solid", start_color="DBEAFE")
	for cell in sheet[1]:
		cell.font = Font(bold=True)
		cell.fill = header_fill

	for row in rows:
		sheet.append(row)

	for col in sheet.columns:
		max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
		sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

	buffer = io.BytesIO()
	wb.save(buffer)

	frappe.response["filename"] = f"{filename_base}.xlsx"
	frappe.response["filecontent"] = buffer.getvalue()
	frappe.response["type"] = "download"
