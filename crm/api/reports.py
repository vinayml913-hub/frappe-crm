"""
Reports API — Deal Status / Trends / Target-Achievement / Leaderboard / Export
================================================================================

Companion module to ``crm/api/revenue.py`` (Revenue Analytics). This module
covers deal-COUNT based reporting (won/lost/open deal counts, leaderboard,
target-vs-achievement by deal count) while ``revenue.py`` covers
deal-VALUE based reporting. Both share the same access model, the same
``CRM Revenue Target`` doctype, and the same Won/Lost/Open classification.

Deal status classification
---------------------------
``CRM Deal Status.type`` is one of: Open, Ongoing, On Hold, Won, Lost
(see crm/install.py add_default_deal_statuses). For reporting purposes:

  - "Won"  -> type == "Won"
  - "Lost" -> type == "Lost"
  - "Open" -> type in ("Open", "Ongoing", "On Hold")  i.e. everything still
              in progress. This matches how the existing Dashboard treats
              "ongoing" deals (see get_avg_ongoing_deal_value in dashboard.py
              which also groups Open+Ongoing+On Hold together via
              `Status.type.notin(["Won", "Lost"])`).

Revenue figure included in KPIs/leaderboard here = Deal.gross_profit
on Won deals only, identical formula to revenue.py, for consistency.
"""

import json

import frappe
from frappe import _
from frappe.query_builder import Case, DocType
from frappe.query_builder.functions import Sum, Count, IfNull, DateFormat
from frappe.utils import getdate, get_first_day, get_last_day, add_months, add_days, nowdate, flt

from crm.api.revenue import (
    _resolve_scope,
    _is_admin,
    _bulk_user_names,
    _get_target_for_range,
    _revenue_expr,
    _deal_team_condition,
)
from crm.api.revenue_target import get_financial_quarter_bounds, get_financial_quarter_for_date


# ─────────────────────────────────────────────────────────────────────────
#  Date range helpers (quick filters)
# ─────────────────────────────────────────────────────────────────────────

QUICK_RANGES = {
    "last_30_days": 30,
}

# Quarter quick-filter keys map to the company's financial quarters
# (Q1 = Mar-May, Q2 = Jun-Aug, Q3 = Sep-Nov, Q4 = Dec-Feb), resolved
# against the fiscal year the *current date* belongs to for that quarter.
QUARTER_KEYS = {"q1": "Q1", "q2": "Q2", "q3": "Q3", "q4": "Q4"}


def _resolve_quick_range(range_key):
    """Translate a quick-filter key into (from_date, to_date)."""
    today = getdate(nowdate())

    if range_key in QUARTER_KEYS:
        quarter = QUARTER_KEYS[range_key]
        _current_quarter, current_quarter_year = get_financial_quarter_for_date(today)
        from_date, to_date = get_financial_quarter_bounds(quarter, current_quarter_year)
        return from_date, to_date

    days = QUICK_RANGES.get(range_key)
    if not days:
        return None, None
    return add_days(today, -days), today


def _default_date_range(from_date, to_date, quick_range=None):
    """
    Resolution order: explicit fromDate/toDate > quick_range key > default
    (last 30 days), matching the spec's quick-filter list.
    """
    if from_date or to_date:
        to_date = getdate(to_date) if to_date else getdate(nowdate())
        from_date = getdate(from_date) if from_date else None
        return from_date, to_date

    if quick_range:
        f, t = _resolve_quick_range(quick_range)
        return f, (t or getdate(nowdate()))

    f, t = _resolve_quick_range("last_30_days")
    return f, t


# ─────────────────────────────────────────────────────────────────────────
#  Core query builder
# ─────────────────────────────────────────────────────────────────────────

def _deal_base_query(from_date, to_date, user=None, date_field="creation"):
    """
    Base query over CRM Deal joined to CRM Deal Status, filtered by a date
    range on the given date field (default: creation date, so "Total Deals"
    reflects deals *created* in range — Won/Lost counts separately filter
    on closed_date, see _won_lost_query below).

    from_date may be None (meaning "Ever" / no lower bound).
    """
    Deal = DocType("CRM Deal")
    Status = DocType("CRM Deal Status")

    field = getattr(Deal, date_field)
    to_date_plus_one = add_days(to_date, 1)

    query = (
        frappe.qb.from_(Deal)
        .join(Status)
        .on(Deal.status == Status.name)
        .where(field < to_date_plus_one)
    )
    if from_date:
        query = query.where(field >= from_date)
    if user:
        query = query.where(_deal_team_condition(Deal, user))

    return query, Deal, Status


def _status_case(Status):
    """A CASE expression collapsing Status.type into Won / Lost / Open."""
    return (
        Case()
        .when(Status.type == "Won", "Won")
        .when(Status.type == "Lost", "Lost")
        .else_("Open")
    )


# ─────────────────────────────────────────────────────────────────────────
#  1. GET /reports/dashboard  ->  crm.api.reports.get_dashboard_kpis
# ─────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_dashboard_kpis(fromDate=None, toDate=None, userId=None, quickRange=None):
    """
    KPI cards: Total / Won / Lost / Open deal counts, Total Target,
    Achieved Target, Achievement %, Revenue Generated.

    "Total/Won/Lost/Open Deals" are counted by deal *creation* date inside
    the range (so the cards answer "how many deals were created in this
    window, and what happened to them"). Revenue is still anchored on
    closed_date for Won deals, matching revenue.py.
    """
    from_date, to_date = _default_date_range(fromDate, toDate, quickRange)
    scoped_user = _resolve_scope(userId)

    query, Deal, Status = _deal_base_query(from_date, to_date, scoped_user, date_field="creation")
    status_expr = _status_case(Status)

    rows = (
        query.select(
            status_expr.as_("bucket"),
            Count(Deal.name).as_("count"),
        )
        .groupby(status_expr)
        .run(as_dict=True)
    )

    counts = {"Won": 0, "Lost": 0, "Open": 0}
    for r in rows:
        counts[r["bucket"]] = int(r["count"] or 0)
    total_deals = sum(counts.values())

    # Revenue generated — Won deals only, by closed_date in range (same
    # convention as revenue.py so the number matches the Revenue tab).
    revenue_query, RDeal, RStatus = _deal_base_query(from_date, to_date, scoped_user, date_field="closed_date")
    revenue_query = revenue_query.where(RStatus.type == "Won").where(RDeal.closed_date.isnotnull())
    revenue_result = revenue_query.select(Sum(_revenue_expr(RDeal)).as_("revenue")).run(as_dict=True)
    revenue_generated = flt(revenue_result[0].get("revenue") or 0) if revenue_result else 0

    # Target / Achievement — only meaningful for a single user or as a
    # company-wide sum across all users with a target in this period.
    total_target, achieved_target = _target_and_achievement(from_date, to_date, scoped_user)
    achievement_pct = (achieved_target / total_target * 100) if total_target else None

    return {
        "scope": "admin" if _is_admin() else "employee",
        "from_date": str(from_date) if from_date else None,
        "to_date": str(to_date),
        "cards": {
            "total_deals": total_deals,
            "won_deals": counts["Won"],
            "lost_deals": counts["Lost"],
            "open_deals": counts["Open"],
            "total_target": total_target,
            "achieved_target": achieved_target,
            "achievement_pct": achievement_pct,
            "revenue_generated": revenue_generated,
        },
    }


def _target_and_achievement(from_date, to_date, user):
    """
    total_target = sum of CRM Revenue Target rows overlapping the range
                    (for the given user, or all users if user is None)
    achieved_target = Won-deal revenue in the same range (closed_date),
                       scoped the same way.
    """
    if user:
        target = _get_target_for_range(user, from_date or get_first_day(to_date), to_date)
        achieved = _won_revenue_in_range(from_date, to_date, user)
        return target, achieved

    # Company-wide: sum target + achieved across every user who has deals
    # or a target in range.
    Deal = DocType("CRM Deal")
    Status = DocType("CRM Deal Status")
    query, _deal, _status = _deal_base_query(from_date, to_date, None, date_field="closed_date")
    query = query.where(Status.type == "Won").where(Deal.closed_date.isnotnull())
    owners = query.select(Deal.deal_owner).distinct().run(as_dict=True)
    owner_names = {o["deal_owner"] for o in owners if o["deal_owner"]}

    target_owners = frappe.get_all(
        "CRM Revenue Target",
        filters={"year": ["in", list(range(getdate(from_date or to_date).year, getdate(to_date).year + 1))]},
        fields=["user"],
        distinct=True,
        ignore_permissions=True,
    )
    owner_names |= {t["user"] for t in target_owners if t["user"]}

    total_target = 0
    total_achieved = 0
    for owner in owner_names:
        total_target += _get_target_for_range(owner, from_date or get_first_day(to_date), to_date) or 0
        total_achieved += _won_revenue_in_range(from_date, to_date, owner)

    return total_target, total_achieved


def _won_revenue_in_range(from_date, to_date, user):
    query, Deal, Status = _deal_base_query(from_date, to_date, user, date_field="closed_date")
    query = query.where(Status.type == "Won").where(Deal.closed_date.isnotnull())
    result = query.select(Sum(_revenue_expr(Deal)).as_("revenue")).run(as_dict=True)
    return flt(result[0].get("revenue") or 0) if result else 0


# ─────────────────────────────────────────────────────────────────────────
#  2. GET /reports/deal-status  ->  crm.api.reports.get_deal_status
#     (feeds both the Donut and Pie chart — same data, two renderings)
# ─────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_deal_status(fromDate=None, toDate=None, userId=None, quickRange=None):
    from_date, to_date = _default_date_range(fromDate, toDate, quickRange)
    scoped_user = _resolve_scope(userId)

    query, Deal, Status = _deal_base_query(from_date, to_date, scoped_user, date_field="creation")
    status_expr = _status_case(Status)

    rows = (
        query.select(status_expr.as_("bucket"), Count(Deal.name).as_("count"))
        .groupby(status_expr)
        .run(as_dict=True)
    )

    counts = {"Won": 0, "Lost": 0, "Open": 0}
    for r in rows:
        counts[r["bucket"]] = int(r["count"] or 0)
    total = sum(counts.values())

    # Matches the project's existing status colors (crm/install.py: Won=green, Lost=red)
    colors = {"Won": "green", "Lost": "red", "Open": "gray"}

    data = [
        {
            "status": k,
            "count": v,
            "percentage": (v / total * 100) if total else 0,
            "color": colors[k],
        }
        for k, v in counts.items()
    ]

    return {
        "from_date": str(from_date) if from_date else None,
        "to_date": str(to_date),
        "total": total,
        "data": data,
        # Ready-to-render <DonutChart> config (frappe-ui)
        "donut_chart": {
            "data": [{"status": d["status"], "count": d["count"]} for d in data],
            "title": _("Deal Status Distribution"),
            "categoryColumn": "status",
            "valueColumn": "count",
        },
    }


# ─────────────────────────────────────────────────────────────────────────
#  3. GET /reports/monthly-trends  ->  crm.api.reports.get_monthly_trends
# ─────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_monthly_trends(fromDate=None, toDate=None, userId=None, quickRange=None):
    """
    Line chart: Won / Lost / Open deal counts per month, bucketed by
    deal creation date. Always returns a complete chronological series
    (zero-filled), same approach as revenue.py's trend endpoint.

    "Ever" / no lower bound falls back to the last 12 months so the chart
    has a sane, bounded x-axis instead of running back to year 1.
    """
    from_date, to_date = _default_date_range(fromDate, toDate, quickRange)
    if not from_date:
        from_date = add_months(get_first_day(to_date), -11)

    scoped_user = _resolve_scope(userId)
    query, Deal, Status = _deal_base_query(from_date, to_date, scoped_user, date_field="creation")

    rows = (
        query.select(
            DateFormat(Deal.creation, "%Y-%m").as_("period"),
            Sum(Case().when(Status.type == "Won", 1).else_(0)).as_("won"),
            Sum(Case().when(Status.type == "Lost", 1).else_(0)).as_("lost"),
            Sum(Case().when(Status.type.notin(["Won", "Lost"]), 1).else_(0)).as_("open"),
        )
        .groupby(DateFormat(Deal.creation, "%Y-%m"))
        .run(as_dict=True)
    )
    by_period = {r["period"]: r for r in rows}

    buckets = []
    cursor = get_first_day(from_date)
    end_marker = get_first_day(to_date)
    while cursor <= end_marker:
        key = cursor.strftime("%Y-%m")
        row = by_period.get(key, {})
        buckets.append({
            "period": key,
            "label": cursor.strftime("%b %Y"),
            "won": int(row.get("won") or 0),
            "lost": int(row.get("lost") or 0),
            "open": int(row.get("open") or 0),
        })
        cursor = add_months(cursor, 1)

    return {
        "from_date": str(from_date),
        "to_date": str(to_date),
        "data": buckets,
        "chart": {
            "data": buckets,
            "title": _("Monthly Deal Trends"),
            "xAxis": {"title": _("Month"), "key": "label", "type": "category"},
            "yAxis": {"title": _("Number of Deals")},
            "series": [
                {"name": "won", "type": "line", "showDataPoints": True},
                {"name": "lost", "type": "line", "showDataPoints": True},
                {"name": "open", "type": "line", "showDataPoints": True},
            ],
        },
    }


# ─────────────────────────────────────────────────────────────────────────
#  4. GET /reports/target-achievement -> crm.api.reports.get_target_achievement
# ─────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_target_achievement(fromDate=None, toDate=None, userId=None, quickRange=None):
    """
    Bar chart: Assigned Target vs Achieved (revenue-based), per employee
    plus an Overall Company row. Thin wrapper that reuses the exact same
    target-resolution + Won-revenue logic as the KPI cards, just broken
    out per employee for charting.
    """
    from_date, to_date = _default_date_range(fromDate, toDate, quickRange)
    scoped_user = _resolve_scope(userId)

    if scoped_user:
        users_to_show = [scoped_user]
    else:
        Deal = DocType("CRM Deal")
        Status = DocType("CRM Deal Status")
        query, _deal, _status = _deal_base_query(from_date, to_date, None, date_field="closed_date")
        query = query.where(Status.type == "Won").where(Deal.closed_date.isnotnull())
        owners = query.select(Deal.deal_owner).distinct().run(as_dict=True)
        users_to_show = sorted({o["deal_owner"] for o in owners if o["deal_owner"]})

        target_owners = frappe.get_all(
            "CRM Revenue Target",
            filters={"year": ["in", list(range(getdate(from_date or to_date).year, getdate(to_date).year + 1))]},
            fields=["user"],
            distinct=True,
            ignore_permissions=True,
        )
        for t in target_owners:
            if t["user"] and t["user"] not in users_to_show:
                users_to_show.append(t["user"])

    user_names = _bulk_user_names(users_to_show)
    results = []
    total_target = 0
    total_achieved = 0

    for u in users_to_show:
        target = _get_target_for_range(u, from_date or get_first_day(to_date), to_date) or 0
        achieved = _won_revenue_in_range(from_date, to_date, u)
        total_target += target
        total_achieved += achieved
        results.append({
            "user": u,
            "employee_name": user_names.get(u, u),
            "target_amount": target,
            "achieved_amount": achieved,
            "achievement_pct": (achieved / target * 100) if target else None,
        })

    results.sort(key=lambda r: r["achieved_amount"], reverse=True)

    overall = {
        "user": None,
        "employee_name": _("Overall Company"),
        "target_amount": total_target,
        "achieved_amount": total_achieved,
        "achievement_pct": (total_achieved / total_target * 100) if total_target else None,
    }

    chart_rows = [{"employee_name": r["employee_name"], "target": r["target_amount"], "achieved": r["achieved_amount"]} for r in results]

    return {
        "from_date": str(from_date) if from_date else None,
        "to_date": str(to_date),
        "overall": overall,
        "data": results,
        "chart": {
            "data": chart_rows,
            "title": _("Target vs Achievement"),
            "xAxis": {"title": _("Employee"), "key": "employee_name", "type": "category"},
            "yAxis": {"title": _("Amount")},
            "series": [
                {"name": "target", "type": "bar"},
                {"name": "achieved", "type": "bar"},
            ],
        },
    }


# ─────────────────────────────────────────────────────────────────────────
#  5. GET /reports/leaderboard  ->  crm.api.reports.get_leaderboard
# ─────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def get_leaderboard(fromDate=None, toDate=None, userId=None, quickRange=None, limit=50):
    """
    Leaderboard: Rank, Employee, Total/Won/Lost/Open Deals, Target,
    Achieved, Achievement %, Revenue. Sorted by Achievement % descending
    (per spec) — employees with no target set sort to the bottom (None
    achievement can't be ranked against a percentage).
    """
    from_date, to_date = _default_date_range(fromDate, toDate, quickRange)
    scoped_user = _resolve_scope(userId)
    limit = int(limit) if limit else 50

    query, Deal, Status = _deal_base_query(from_date, to_date, scoped_user, date_field="creation")
    status_expr = _status_case(Status)

    rows = (
        query.select(
            Deal.deal_owner.as_("user"),
            status_expr.as_("bucket"),
            Count(Deal.name).as_("count"),
        )
        .groupby(Deal.deal_owner, status_expr)
        .run(as_dict=True)
    )

    by_user = {}
    for r in rows:
        u = r["user"]
        if u not in by_user:
            by_user[u] = {"total": 0, "Won": 0, "Lost": 0, "Open": 0}
        by_user[u][r["bucket"]] = int(r["count"] or 0)
        by_user[u]["total"] += int(r["count"] or 0)

    user_names = _bulk_user_names(list(by_user.keys()))

    results = []
    for u, counts in by_user.items():
        target = _get_target_for_range(u, from_date or get_first_day(to_date), to_date) or 0
        revenue = _won_revenue_in_range(from_date, to_date, u)
        achievement_pct = (revenue / target * 100) if target else None

        results.append({
            "user": u,
            "employee_name": user_names.get(u, u),
            "total_deals": counts["total"],
            "won_deals": counts["Won"],
            "lost_deals": counts["Lost"],
            "open_deals": counts["Open"],
            "target_amount": target,
            "achieved_amount": revenue,
            "achievement_pct": achievement_pct,
            "revenue": revenue,
        })

    # Sort by Achievement % descending; entries with no target (None) go last.
    results.sort(key=lambda r: (r["achievement_pct"] is None, -(r["achievement_pct"] or 0)))
    results = results[:limit]

    for idx, r in enumerate(results, start=1):
        r["rank"] = idx

    return {
        "from_date": str(from_date) if from_date else None,
        "to_date": str(to_date),
        "data": results,
    }


# ─────────────────────────────────────────────────────────────────────────
#  6. GET /reports/export  ->  crm.api.reports.export_report
# ─────────────────────────────────────────────────────────────────────────

@frappe.whitelist()
def export_report(fromDate=None, toDate=None, userId=None, quickRange=None, format="csv", source="leaderboard"):
    """
    Streams a file download (CSV or XLSX) of the currently filtered
    leaderboard data using Frappe's standard frappe.response file-download
    mechanism (the same approach Frappe core uses for its own report
    exports — no new download endpoint pattern introduced).

    Called directly via a browser navigation / window.open from the
    frontend (not through `call()`/createResource, since this returns a
    binary file, not JSON) — see ExportButtons.vue.
    """
    leaderboard = get_leaderboard(fromDate, toDate, userId, quickRange, limit=10000)
    rows = leaderboard["data"]

    headers = [
        "Rank", "Employee Name", "Total Deals", "Won Deals", "Lost Deals",
        "Open Deals", "Target", "Achieved", "Achievement %", "Revenue",
    ]

    def row_values(r):
        return [
            r["rank"], r["employee_name"], r["total_deals"], r["won_deals"],
            r["lost_deals"], r["open_deals"], r["target_amount"], r["achieved_amount"],
            f'{r["achievement_pct"]:.1f}' if r["achievement_pct"] is not None else "",
            r["revenue"],
        ]

    filename_base = f"reports-{leaderboard.get('from_date') or 'all'}-to-{leaderboard.get('to_date')}"

    if format == "xlsx":
        _export_xlsx(headers, [row_values(r) for r in rows], filename_base)
    else:
        _export_csv(headers, [row_values(r) for r in rows], filename_base)


def _export_csv(headers, rows, filename_base):
    import csv
    import io

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)

    frappe.response["filename"] = f"{filename_base}.csv"
    frappe.response["filecontent"] = buffer.getvalue()
    frappe.response["type"] = "download"


def _export_xlsx(headers, rows, filename_base):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import io

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Report"

    sheet.append(headers)
    header_fill = PatternFill("solid", start_color="E5E7EB")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        sheet.append(row)

    for col in sheet.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)

    frappe.response["filename"] = f"{filename_base}.xlsx"
    frappe.response["filecontent"] = buffer.getvalue()
    frappe.response["type"] = "download"
